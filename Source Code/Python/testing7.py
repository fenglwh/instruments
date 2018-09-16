from labinstrument.SS.CMW500.CMW500_WIFI.CMW500_WIFI import CMW_WIFI
import openpyxl


if __name__ == '__main__':
    instrument=CMW_WIFI(17)

    print(instrument.uddrate_mode)
    print(instrument.DSSS_rate)
    print(instrument.OFDM_rate)
    print(instrument.OMCS_rate)
    instrument.tx_modulation_format='DSSS'
    tech_2g=1
    tech_5g=0

    if tech_2g:
        wb=openpyxl.Workbook()
        ws=wb.worksheets[0]
        row_count=1
        for standard in ['BSTD','GSTD','GNST']:
            for channel in [1,6,13]:
                instrument.signal_off()
                instrument.standard = standard
                instrument.channel = channel
                if standard=='BSTD':
                    instrument.tx_modulation_format='DSSS'
                    instrument.uddrate_mode='True'
                    instrument.DSSS_rate='DIS,DIS,DIS,MAND'
                    instrument.MFR_control_rate='ENAB,NHT,C11M'
                    instrument.DFR_control_rate='ENAB,NHT,C11M'
                    instrument.write('CONFigure:WLAN:MEAS:ISIGnal:STANdard DSSS')
                elif standard=='GSTD':
                    instrument.tx_modulation_format='OFDM'
                    instrument.uddrate_mode='True'
                    instrument.DSSS_rate='DIS,DIS,DIS,DIS'
                    instrument.OFDM_rate='MAND,DIS,DIS,DIS,DIS,DIS,DIS,DIS'
                    instrument.MFR_control_rate='ENAB,NHT,BR12'
                    instrument.DFR_control_rate='ENAB,NHT,BR12'
                    instrument.write('CONFigure:WLAN:MEAS:ISIGnal:STANdard LOFD')
                elif standard=='GNST':
                    instrument.tx_modulation_format = 'OFDM'
                    instrument.uddrate_mode = 'True'
                    instrument.DSSS_rate = 'DIS,DIS,DIS,DIS'
                    instrument.OFDM_rate = 'OPT,DIS,DIS,DIS,DIS,DIS,DIS,DIS'
                    instrument.OMCS_rate ='SUPP,NOTS,NOTS,NOTS,NOTS,NOTS,NOTS,NOTS'
                    instrument.MFR_control_rate='ENAB,HTM,MCS'
                    instrument.DFR_control_rate='ENAB,HTM,MCS'
                    instrument.write('CONFigure:WLAN:MEAS:ISIGnal:STANdard HTOF')
                instrument.signal_on()
                instrument.packet_generator_ON()
                instrument.wait_for_connect()
                for x in range(3):
                    power=instrument.meas_tx_ping()
                    if power!='NACK':
                        break
                print("{} | {} | {}".format(standard,channel,power))
                ws.cell(row=row_count,column=1).value=str(standard)
                ws.cell(row=row_count+1,column=2).value=str(channel)
                ws.cell(row=row_count+2,column=3).value=str(power)
                row_count+=3
        wb.save('2G.xlsx')
        wb.close()

    #5g
    if tech_5g:
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        row_count = 1
        for standard in ['ASTD', 'ANST']:
            for channel in [36,44,48,52,60,64,100,120,140,149,157,161,165]:
                instrument.signal_off()
                instrument.standard = standard
                instrument.channel = channel
                if standard=='ASTD':
                    instrument.tx_modulation_format = 'OFDM'
                    instrument.uddrate_mode = 'True'
                    instrument.DSSS_rate = 'DIS,DIS,DIS,DIS'
                    instrument.OFDM_rate = 'MAND,DIS,DIS,DIS,DIS,DIS,DIS,DIS'
                    instrument.MFR_control_rate = 'ENAB,NHT,BR12'
                    instrument.DFR_control_rate = 'ENAB,NHT,BR12'
                    instrument.write('CONFigure:WLAN:MEAS:ISIGnal:STANdard LOFD')
                elif standard=='ANST':
                    instrument.tx_modulation_format = 'OFDM'
                    instrument.uddrate_mode = 'True'
                    instrument.DSSS_rate = 'DIS,DIS,DIS,DIS'
                    instrument.OFDM_rate = 'OPT,DIS,DIS,DIS,DIS,DIS,DIS,DIS'
                    instrument.OMCS_rate = 'SUPP,NOTS,NOTS,NOTS,NOTS,NOTS,NOTS,NOTS'
                    instrument.MFR_control_rate = 'ENAB,HTM,MCS'
                    instrument.DFR_control_rate = 'ENAB,HTM,MCS'
                    instrument.write('CONFigure:WLAN:MEAS:ISIGnal:STANdard HTOF')
                instrument.signal_on()
                instrument.packet_generator_ON()
                instrument.wait_for_connect()
                for x in range(3):
                    power=instrument.meas_tx_ping()
                    if power!='NACK':
                        break
                print("{} | {} | {}".format(standard, channel,power ))
                ws.cell(row=row_count,column=1).value=str(standard)
                ws.cell(row=row_count,column=2).value=str(channel)
                ws.cell(row=row_count,column=3).value=str(power)
                row_count+=1
        wb.save('5G.xlsx')
        wb.close()

